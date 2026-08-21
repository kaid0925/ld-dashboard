import json, os, glob, re
from collections import defaultdict
os.chdir(os.path.dirname(os.path.abspath(__file__)))
BASE=os.path.dirname(os.path.abspath(__file__))
DI=os.path.join(BASE,'drive_in'); STATE=os.path.join(BASE,'dash_data.json')
EXCLUDE=['엘마운트 부품','설치지원']; TRACK=['AP','LDL','PMN','SS','SPL']
BRANDS=['AP','LDL','PMN','SS','SPL','ETC']
def brand(name):
    n=(name or ''); u=n.upper()
    if any(x in n for x in EXCLUDE): return 'ETC'
    if ('엘마운트' in n) or ('엘디엘마운트' in n) or ('엘디엘 마운트' in n) or ('APL-' in u): return 'LDL'
    if ('포미니' in n) or ('미피' in n): return 'PMN'
    if '신성' in n: return 'SS'
    if ('쏘플링' in n) or ('SPL-' in u): return 'SPL'
    if ('애니포트' in n) or ('ANYPORT' in u): return 'AP'
    return 'ETC'
def model_of(name):
    for m in re.finditer(r'\[([^\[\]]*)\]', name or ''):
        c=m.group(1)
        if 'APL-' in c.upper() or 'AP-' in c.upper() or 'SPL-' in c.upper(): return re.sub(r'★','',c).strip()
    u=(name or '').upper()
    m=re.search(r'(APL-[A-Z0-9][A-Z0-9\-]*)',u) or re.search(r'(AP-[A-Z0-9][A-Z0-9\-]*)',u) or re.search(r'(SPL-[A-Z0-9][A-Z0-9\-]*)',u)
    return m.group(1).strip() if m else None
def clean(name): return re.sub(r'\s+',' ',(name or '').replace('**스마일배송**','').strip())
def market_channel(g):
    for p,disp in [('옥션','옥션'),('G마켓','G마켓'),('11번가','11번가'),('스마트스토어','스마트스토어'),('공통엑셀양식','쿠팡FC/VF'),('쿠팡','쿠팡'),('롯데온','롯데온'),('알리익스프레스','알리익스프레스')]:
        if g.startswith(p): return disp
    return {'이지웰':'이지웰','삼성카드(복지)':'삼성카드VIP몰','오늘의집':'오늘의집','카카오선물하기':'카카오선물하기'}.get(g)
SITE=[('삼성카드(쇼핑)','삼성카드(쇼핑)'),('삼성카드(복지)','삼성카드VIP몰'),('제이슨딜','제이슨딜'),('플록(flock)-원룸만들기','원룸만들기'),('비버커뮤니케이션','비버커뮤니케이션'),('유니브','비누커머스'),('아트박스','아트박스'),('지그재그','지그재그'),('롯데하이마트','롯데하이마트'),('에스에스지닷컴','SSG'),('토스쇼핑','토스쇼핑'),('넛지헬스케어','넛지'),('무신사','무신사'),('에이블리','에이블리'),('GS SHOP','GS SHOP'),('LG전자(LG복지몰)','LG복지몰')]
def site_channel(v):
    for k,disp in SITE:
        if k in v: return disp
    return None
LUMP_MONTHS=set()
def pkey_of(datev):
    d=str(datev)[:10]; mo=d[5:7]; dd=d[8:10]
    if not (mo.isdigit() and dd.isdigit()): return None
    if mo in LUMP_MONTHS: return str(int(mo))
    return f'{int(mo)}-{dd}'
def plabel(pk):
    if '-' in pk: mo,dd=pk.split('-'); return f'{int(mo)}/{int(dd)}'
    return f'{int(pk)}월'
def pk_sort(pk):
    if '-' in pk: mo,dd=pk.split('-'); return (int(mo),int(dd))
    return (int(pk),0)
MONTHEND={'2026-05-31':'5월말','2026-06-30':'6월말','2026-07-31':'7월말'}
import openpyxl
st=json.load(open(STATE,encoding='utf-8'))
existing_pk=set(p['k'] for p in st['periods'])
existing_snap=set(s['k'] for s in st['snaps'])
model_by={m['key']:m for m in st['models']}

# ===== NEW SALES =====
sales_files=sorted(glob.glob(f'{DI}/매출이익리스트*.xlsx'))
_sm=set()
for _p in st['periods']:
    _k=_p['k']; _sm.add(str(int(_k.split('-')[0])).zfill(2) if '-' in _k else str(int(_k)).zfill(2))
for _f in sales_files:
    _wb=openpyxl.load_workbook(_f,read_only=True,data_only=True); _ws=_wb[_wb.sheetnames[0]]
    for _r in _ws.iter_rows(min_row=3,values_only=True):
        _d=str(_r[0])[:10]
        if len(_d)==10 and _d[:4]=='2026' and _d[5:7].isdigit(): _sm.add(_d[5:7])
    _wb.close()
LUMP_MONTHS={m for m in _sm if _sm and m<max(_sm)} if _sm else set()
date_best={}
for f in sales_files:
    m=re.search(r'~2026-(\d\d)-(\d\d)\)',f); end=(int(m.group(1)),int(m.group(2))) if m else (99,99)
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=3,values_only=True):
        d=str(r[0])[:10]
        if len(d)==10 and d[:4]=='2026' and d[5:7] not in LUMP_MONTHS:
            if d not in date_best or end>date_best[d][0]: date_best[d]=(end,f)
    wb.close()
new_ch=defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:[0,0,0])))
new_mdl=defaultdict(lambda:{'name':'','cat':'','p':defaultdict(int),'ch':{}}); new_pks=set()
for f in sales_files:
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=3,values_only=True):
        d=str(r[0])[:10]
        if len(d)==10 and d[5:7] not in LUMP_MONTHS and date_best.get(d,(None,None))[1]!=f: continue
        pk=pkey_of(r[0])
        if not pk or pk in existing_pk: continue
        new_pks.add(pk)
        g=(r[3] or '').strip(); v=(r[7] or '').strip()
        ch=market_channel(g) if g else site_channel(v)
        if not ch: continue
        nm=(r[31] or '').strip(); b=brand(nm)
        try:qty=int(r[42] or 0)
        except:qty=0
        try:amt=int(r[44] or 0)
        except:amt=0
        try:prof=int(r[53] or 0)
        except:prof=0
        c=new_ch[ch][pk][b];c[0]+=qty;c[1]+=amt;c[2]+=prof
        if b in TRACK:
            key=b+'|'+(model_of(nm) or clean(nm)); e=new_mdl[key];e['p'][pk]+=qty
            ce=e['ch'].setdefault(ch,[0,0]);ce[0]+=qty;ce[1]+=amt
            if not e['name']:e['name']=clean(nm)
            if not e['cat']:e['cat']=(r[27] or '').strip()
    wb.close()
# merge new periods
for pk in sorted(new_pks,key=pk_sort):
    st['periods'].append({'k':pk,'l':plabel(pk),'t':('m' if '-' not in pk else 'd')})
    for c in st['channels']:
        c['p'][pk]={b:new_ch[c['name']][pk][b] for b in BRANDS}
# new models (keys not in state)
for key,e in new_mdl.items():
    if key in model_by: continue
    mk=key.split('|')[1]
    nm={'key':key,'brand':key.split('|')[0],'model':(mk if re.match(r'(AP|SPL)',mk) else ''),'name':e['name'],'cat':e['cat'],
        'p':{p['k']:0 for p in st['periods']},'total':0,'avg':0,'stock':0,'stock_s':{s['k']:0 for s in st['snaps']},'stock_amt':0,'stock_amt_s':{s['k']:0 for s in st['snaps']},'ch':{}}
    st['models'].append(nm); model_by[key]=nm
# fill model p for new pks
for m in st['models']:
    m.setdefault('ch',{})
    nc=new_mdl.get(m['key'],{}).get('ch',{})
    for ch,vv in nc.items():
        cc=m['ch'].setdefault(ch,[0,0]); cc[0]+=vv[0]; cc[1]+=vv[1]
    for pk in new_pks:
        m['p'][pk]=new_mdl.get(m['key'],{'p':{}})['p'].get(pk,0) if m['key'] in new_mdl else m['p'].get(pk,0)
        if pk not in m['p']: m['p'][pk]=0
# ensure every model has all pks
allpk=[p['k'] for p in st['periods']]
for m in st['models']:
    for pk in allpk:
        m['p'].setdefault(pk,0)

# ===== NEW INVENTORY SNAPSHOTS =====
snap_seen={}
for f in sorted(glob.glob(f'{DI}/재고현황*.xlsx')):
    m=re.search(r'\((\d{4}-\d\d-\d\d)\)',f)
    if m: snap_seen[m.group(1)]=f
for S in sorted(snap_seen):
    if S in existing_snap: continue
    fn=snap_seen[S]
    wb=openpyxl.load_workbook(fn,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    items=[]; bs={b:[0,0,0] for b in TRACK}; sm=defaultdict(lambda:[0,0])
    for r in ws.iter_rows(min_row=2,values_only=True):
        nm=(r[7] or '').strip(); b=brand(nm)
        if b not in TRACK: continue
        try:qty=int(r[17] or 0)
        except:qty=0
        try:amt=int(r[20] or 0)
        except:amt=0
        unit=int(r[19] or 0)
        items.append({'brand':b,'name':clean(nm),'code':str(r[0] or ''),'cat':(r[27] or '').strip(),'qty':qty,'unit':unit,'amt':amt})
        bs[b][0]+=1;bs[b][1]+=qty;bs[b][2]+=amt
        key=b+'|'+(model_of(nm) or clean(nm)); sm[key][0]+=qty; sm[key][1]+=amt
    items.sort(key=lambda x:-x['amt'])
    wb.close()
    st['snaps'].append({'k':S,'l':MONTHEND.get(S, f'{int(S[5:7])}/{int(S[8:10])}')})
    st['bstock_by_snap'][S]=bs; st['inv_by_snap'][S]=items
    for m in st['models']:
        m['stock_s'][S]=sm.get(m['key'],[0,0])[0]
        m.setdefault('stock_amt_s',{})[S]=sm.get(m['key'],[0,0])[1]

# ===== finalize =====

def collapse_completed(st):
    from collections import defaultdict as _dd
    grp=_dd(list)
    for p in st['periods']:
        k=p['k']
        if '-' in k:
            mo=str(int(k.split('-')[0])).zfill(2)
            if mo in LUMP_MONTHS: grp[mo].append(k)
    for mo,dks in grp.items():
        lk=str(int(mo))
        for c in st['channels']:
            agg={b:[0,0,0] for b in ['AP','LDL','PMN','SS','SPL','ETC']}
            for dk in dks:
                cell=c['p'].get(dk,{})
                for b in agg:
                    v=cell.get(b,[0,0,0]); agg[b][0]+=v[0]; agg[b][1]+=v[1]; agg[b][2]+=v[2]
                c['p'].pop(dk,None)
            base=c['p'].get(lk,{b:[0,0,0] for b in agg})
            for b in agg:
                bb=base.get(b,[0,0,0]); agg[b][0]+=bb[0]; agg[b][1]+=bb[1]; agg[b][2]+=bb[2]
            c['p'][lk]=agg
        for m in st['models']:
            t=0
            for dk in dks: t+=m['p'].pop(dk,0)
            m['p'][lk]=m['p'].get(lk,0)+t
        st['periods']=[p for p in st['periods'] if p['k'] not in dks]
        if not any(p['k']==lk for p in st['periods']):
            st['periods'].append({'k':lk,'l':f'{int(mo)}월','t':'m'})
collapse_completed(st)
st['periods'].sort(key=lambda p:pk_sort(p['k']))
st['snaps'].sort(key=lambda s:s['k'])
st['latest_snap']=st['snaps'][-1]['k']
st['day_keys']=[p['k'] for p in st['periods'] if '-' in p['k']]
LATEST=st['latest_snap']; lump=[p['k'] for p in st['periods'] if '-' not in p['k']][:3]
for m in st['models']:
    m['total']=sum(m['p'].values())
    m['avg']=round(sum(m['p'].get(k,0) for k in lump)/len(lump),1) if lump else 0
    m['stock']=m['stock_s'].get(LATEST,0)
    m['stock_amt']=m.get('stock_amt_s',{}).get(LATEST,0)
st['models'].sort(key=lambda x:-x['total'])
json.dump(st,open(STATE,'w',encoding='utf-8'),ensure_ascii=False)
print('신규 기간:',sorted(new_pks,key=pk_sort),'| 신규 스냅샷:',[S for S in sorted(snap_seen) if S not in existing_snap])
print('periods:',[p['l'] for p in st['periods']],'| latest:',st['latest_snap'],'| models:',len(st['models']))
