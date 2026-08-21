import openpyxl, glob, re, json
from collections import defaultdict
import os
BASE=os.path.dirname(os.path.abspath(__file__))
DI=os.path.join(BASE,'drive_in')
OUT=os.path.join(BASE,'dash_data.json')
EXCLUDE=['엘마운트 부품','설치지원']
def brand(name):
    n=(name or ''); u=n.upper()
    if any(x in n for x in EXCLUDE): return 'ETC'
    if ('엘마운트' in n) or ('엘디엘마운트' in n) or ('엘디엘 마운트' in n) or ('APL-' in u): return 'LDL'
    if ('포미니' in n) or ('미피' in n): return 'PMN'
    if '신성' in n: return 'SS'
    if ('쏘플링' in n) or ('SPL-' in u): return 'SPL'
    if ('애니포트' in n) or ('ANYPORT' in u): return 'AP'
    return 'ETC'
TRACK=['AP','LDL','PMN','SS','SPL']
def model_of(name):
    for m in re.finditer(r'\[([^\[\]]*)\]', name or ''):
        c=m.group(1)
        if 'APL-' in c.upper() or 'AP-' in c.upper() or 'SPL-' in c.upper(): return re.sub(r'★','',c).strip()
    u=(name or '').upper()
    m=re.search(r'(APL-[A-Z0-9][A-Z0-9\-]*)',u) or re.search(r'(AP-[A-Z0-9][A-Z0-9\-]*)',u) or re.search(r'(SPL-[A-Z0-9][A-Z0-9\-]*)',u)
    return m.group(1).strip() if m else None
def clean(name): return re.sub(r'\s+',' ',(name or '').replace('**스마일배송**','').strip())
def market_channel(g):
    for p,disp in [('옥션','옥션'),('G마켓','G마켓'),('11번가','11번가'),('스마트스토어','스마트스토어'),('공통엑셀양식','쿠팡FC/VF'),('쿠팡','쿠팡'),('롯데온','롯데온'),('알리익스프레스','알리익스프레스')]:
        if g.startswith(p): return disp
    return {'이지웰':'이지웰','삼성카드(복지)':'삼성카드VIP몰','오늘의집':'오늘의집','카카오선물하기':'카카오선물하기'}.get(g)
SITE=[('삼성카드(쇼핑)','삼성카드(쇼핑)'),('삼성카드(복지)','삼성카드VIP몰'),('제이슨딜','제이슨딜'),('플록(flock)-원룸만들기','원룸만들기'),('비버커뮤니케이션','비버커뮤니케이션'),('유니브','비누커머스'),('아트박스','아트박스'),('지그재그','지그재그'),('롯데하이마트','롯데하이마트'),('에스에스지닷컴','SSG'),('토스쇼핑','토스쇼핑'),('넛지헬스케어','넛지'),('무신사','무신사'),('에이블리','에이블리'),('GS SHOP','GS SHOP'),('LG전자(LG복지몰)','LG복지몰')]
def site_channel(v):
    for k,disp in SITE:
        if k in v: return disp
    return None
order=['옥션','G마켓','11번가','스마트스토어','쿠팡','쿠팡FC/VF','카카오선물하기','롯데온','이지웰','삼성카드VIP몰','삼성카드(쇼핑)','오늘의집','알리익스프레스','제이슨딜','원룸만들기','비버커뮤니케이션','비누커머스','아트박스','지그재그','롯데하이마트','SSG','토스쇼핑','넛지','무신사','에이블리','GS SHOP','LG복지몰','텐바이텐']
LUMP_MONTHS=set()  # 자동판정(아래에서 최신달 제외 나머지)
def pkey_of(datev):
    d=str(datev)[:10]; mo=d[5:7]; dd=d[8:10]
    if not (mo.isdigit() and dd.isdigit()): return None
    if mo in LUMP_MONTHS: return str(int(mo))
    return f'{int(mo)}-{dd}'
def plabel(pk):
    if '-' in pk: mo,dd=pk.split('-'); return f'{int(mo)}/{int(dd)}'
    return f'{int(pk)}월'
sales_files=sorted(glob.glob(f'{DI}/매출이익리스트*.xlsx'))
_months=set()
for _f in sales_files:
    _wb=openpyxl.load_workbook(_f,read_only=True,data_only=True); _ws=_wb[_wb.sheetnames[0]]
    for _r in _ws.iter_rows(min_row=3,values_only=True):
        _d=str(_r[0])[:10]
        if len(_d)==10 and _d[:4]=='2026' and _d[5:7].isdigit(): _months.add(_d[5:7])
    _wb.close()
if _months:
    _latest=max(_months); LUMP_MONTHS={m for m in _months if m<_latest}
date_best={}
for f in sales_files:
    m=re.search(r'~2026-(\d\d)-(\d\d)\)',f); end=(int(m.group(1)),int(m.group(2))) if m else (99,99)
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=3,values_only=True):
        d=str(r[0])[:10]
        if len(d)==10 and d[:4]=='2026' and d[5:7] not in LUMP_MONTHS:
            if d not in date_best or end>date_best[d][0]: date_best[d]=(end,f)
    wb.close()
ch_p=defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:[0,0,0])))
mdl=defaultdict(lambda:{'name':'','cat':'','p':defaultdict(int)}); pkeys=set()
for f in sales_files:
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=3,values_only=True):
        d=str(r[0])[:10]
        if len(d)==10 and d[5:7] not in LUMP_MONTHS and date_best.get(d,(None,None))[1]!=f: continue
        pk=pkey_of(r[0])
        if not pk: continue
        pkeys.add(pk)
        g=(r[3] or '').strip(); v=(r[7] or '').strip()
        ch=market_channel(g) if g else site_channel(v)
        if not ch: continue
        name=(r[31] or '').strip(); b=brand(name)
        try:qty=int(r[42] or 0)
        except:qty=0
        try:amt=int(r[44] or 0)
        except:amt=0
        try:prof=int(r[53] or 0)
        except:prof=0
        c=ch_p[ch][pk][b];c[0]+=qty;c[1]+=amt;c[2]+=prof
        if b in TRACK:
            key=b+'|'+(model_of(name) or clean(name)); e=mdl[key];e['p'][pk]+=qty
            if not e['name']:e['name']=clean(name)
            if not e['cat']:e['cat']=(r[27] or '').strip()
    wb.close()
MONTHEND={'2026-05-31':'5월말','2026-06-30':'6월말','2026-07-31':'7월말'}
snap_seen={}
for f in sorted(glob.glob(f'{DI}/재고현황*.xlsx')):
    m=re.search(r'\((\d{4}-\d\d-\d\d)\)',f)
    if not m: continue
    dt=m.group(1)
    snap_seen[dt]=f  # 같은 날짜면 나중 파일로 덮어씀(중복 제거)
snap_files=[(snap_seen[dt],dt,MONTHEND.get(dt, f'{int(dt[5:7])}/{int(dt[8:10])}')) for dt in sorted(snap_seen)]
LATEST=snap_files[-1][1]
inv_by_snap={}; bstock_by_snap={}; smodel_by_snap={}
for fn,sk,sl in snap_files:
    wb=openpyxl.load_workbook(fn,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
    items=[]; bs={b:[0,0,0] for b in TRACK}; sm=defaultdict(lambda:[0,0])
    for r in ws.iter_rows(min_row=2,values_only=True):
        name=(r[7] or '').strip(); b=brand(name)
        if b not in TRACK: continue
        try:qty=int(r[17] or 0)
        except:qty=0
        try:amt=int(r[20] or 0)
        except:amt=0
        unit=int(r[19] or 0)
        items.append({'brand':b,'name':clean(name),'code':str(r[0] or ''),'cat':(r[27] or '').strip(),'qty':qty,'unit':unit,'amt':amt})
        bs[b][0]+=1;bs[b][1]+=qty;bs[b][2]+=amt
        key=b+'|'+(model_of(name) or clean(name)); sm[key][0]+=qty; sm[key][1]+=amt
    items.sort(key=lambda x:-x['amt'])
    inv_by_snap[sk]=items; bstock_by_snap[sk]=bs; smodel_by_snap[sk]=dict(sm)
    wb.close()
def pk_sort(pk):
    if '-' in pk: mo,dd=pk.split('-'); return (int(mo),int(dd))
    return (int(pk),0)
allpk=sorted(pkeys,key=pk_sort)
periods=[{'k':pk,'l':plabel(pk),'t':('m' if '-' not in pk else 'd')} for pk in allpk]
allk=[p['k'] for p in periods]
lump=[pk for pk in allk if '-' not in pk][:3]
channels=[{'name':ch,'p':{pk:{b:ch_p[ch][pk][b] for b in['AP','LDL','PMN','SS','SPL','ETC']} for pk in allk}} for ch in order]
snaps=[{'k':sk,'l':sl} for fn,sk,sl in snap_files]
models=[]
for key,e in mdl.items():
    b=key.split('|')[0]; p={pk:e['p'].get(pk,0) for pk in allk}; total=sum(p.values())
    if total<=0: continue
    avg=round(sum(p[k] for k in lump)/len(lump),1) if lump else 0
    stock_s={sk:smodel_by_snap[sk].get(key,[0,0])[0] for fn,sk,sl in snap_files}
    models.append({'key':key,'brand':b,'model':key.split('|')[1] if re.match(r'(AP|SPL)',key.split('|')[1]) else '','name':e['name'],'cat':e['cat'],'p':p,'total':total,'avg':avg,'stock':stock_s[LATEST],'stock_s':stock_s})
models.sort(key=lambda x:-x['total'])
day_keys=[pk for pk in allk if '-' in pk]
out={'periods':periods,'day_keys':day_keys,'channels':channels,'models':models,'snaps':snaps,'latest_snap':LATEST,'inv_by_snap':inv_by_snap,'bstock_by_snap':bstock_by_snap}
json.dump(out,open(OUT,'w',encoding='utf-8'),ensure_ascii=False)
print('periods:',[p['l'] for p in periods])
print('snaps:',[s['l'] for s in snaps],'| latest:',LATEST,'| models:',len(models))
for fn,sk,sl in snap_files: print(f"  {sl}: AP {bstock_by_snap[sk]['AP'][2]:,}  LDL {bstock_by_snap[sk]['LDL'][2]:,}")
